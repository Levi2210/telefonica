from flask import Flask, render_template, request, redirect, url_for, flash
from flask_mysqldb import MySQL
import os
from openpyxl import load_workbook
import re

app = Flask(__name__)

# Configuración de la base de datos
app.config['MYSQL_HOST'] = 'localhost'  
app.config['MYSQL_USER'] = 'root' 
app.config['MYSQL_PASSWORD'] = 'milton'  
app.config['MYSQL_DB'] = 'telefonica_db'

mysql = MySQL(app)

# Configuración secreta para sesiones y mensajes flash
app.secret_key = 'tu_clave_secreta'

# Función para validar el formato del correo electrónico
def es_email_valido(email):
    patron_email = r'^[\w\.-]+@[\w\.-]+\.\w+$'
    return re.match(patron_email, email)

# Función para validar el RUT
def es_rut_valido(rut):
    return len(rut.strip()) > 0

# Ruta para agregar un nuevo cliente
@app.route('/agregar_cliente', methods=['GET', 'POST'])
def agregar_cliente():
    if request.method == 'POST':
        usuario = request.form['usuario']
        estatus = request.form['estatus']
        rut = request.form['rut']
        nombre = request.form['nombre']
        direccion = request.form['direccion']
        email = request.form['email']
        telefono = request.form['telefono']
        tipo_plan = request.form['tipo_plan']
        tipo_cambio = request.form['tipo_cambio']

        # Validar los datos aquí
        if not es_email_valido(email):
            flash("El correo electrónico ingresado no es válido.", "danger")
            return redirect(request.url)
        if not es_rut_valido(rut):
            flash("El RUT ingresado no es válido.", "danger")
            return redirect(request.url)

        cursor = mysql.connection.cursor()
        cursor.execute(
            "INSERT INTO clientes (usuario, estatus, rut, nombre, direccion, email, telefono, tipo_plan, tipo_cambio) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)",
            (usuario, estatus, rut, nombre, direccion, email, telefono, tipo_plan, tipo_cambio)
        )
        mysql.connection.commit()
        cursor.close()

        flash("Cliente agregado exitosamente", "success")
        return redirect(url_for('lista_clientes'))

    return render_template('agregar_cliente.html')

# Ruta para editar un cliente existente
@app.route('/editar_cliente/<int:id>', methods=['GET', 'POST'])
def editar_cliente(id):
    cursor = mysql.connection.cursor()
    if request.method == 'POST':
        # Obtener los datos del formulario de edición
        usuario = request.form['usuario']
        estatus = request.form['estatus']
        rut = request.form['rut']
        nombre = request.form['nombre']
        direccion = request.form['direccion']
        email = request.form['email']
        telefono = request.form['telefono']
        tipo_plan = request.form['tipo_plan']
        tipo_cambio = request.form['tipo_cambio']

        # Validar los datos aquí
        if not es_email_valido(email):
            flash("El correo electrónico ingresado no es válido.", "danger")
            return redirect(request.url)
        if not es_rut_valido(rut):
            flash("El RUT ingresado no es válido.", "danger")
            return redirect(request.url)

        cursor.execute(
            "UPDATE clientes SET usuario=%s, estatus=%s, rut=%s, nombre=%s, direccion=%s, email=%s, telefono=%s, tipo_plan=%s, tipo_cambio=%s WHERE id=%s",
            (usuario, estatus, rut, nombre, direccion, email, telefono, tipo_plan, tipo_cambio, id)
        )
        mysql.connection.commit()
        cursor.close()

        flash("Cliente actualizado exitosamente", "success")
        return redirect(url_for('lista_clientes'))

    cursor.execute("SELECT * FROM clientes WHERE id = %s", (id,))
    cliente = cursor.fetchone()
    cursor.close()

    return render_template('editar_cliente.html', cliente=cliente)

# Ruta para confirmar la eliminación de un cliente
@app.route('/confirmar_eliminar/<int:id>')
def confirmar_eliminar(id):
    cursor = mysql.connection.cursor()
    cursor.execute("SELECT * FROM clientes WHERE id = %s", (id,))
    cliente = cursor.fetchone()
    cursor.close()

    return render_template('confirmar_eliminar.html', cliente=cliente)

# Ruta para eliminar un cliente
@app.route('/eliminar_cliente/<int:id>', methods=['POST'])
def eliminar_cliente(id):
    cursor = mysql.connection.cursor()
    cursor.execute("DELETE FROM clientes WHERE id = %s", (id,))
    mysql.connection.commit()
    cursor.close()

    flash("Cliente eliminado exitosamente", "success")
    return redirect(url_for('lista_clientes'))

# Ruta para listar todos los clientes
@app.route('/lista_clientes')
def lista_clientes():
    cursor = mysql.connection.cursor()
    cursor.execute("SELECT * FROM clientes")
    clientes = cursor.fetchall()
    cursor.close()

    return render_template('lista_clientes.html', clientes=clientes)

# Ruta para cargar datos desde un archivo Excel
@app.route('/cargar_excel', methods=['GET', 'POST'])
def cargar_excel():
    if request.method == 'POST':
        archivo = request.files['archivo_excel']

        if archivo:
            try:
                # Carga el archivo Excel
                archivo_nombre = archivo.filename
                archivo.save(os.path.join('uploads', archivo_nombre))
                archivo_path = os.path.join('uploads', archivo_nombre)
                workbook = load_workbook(archivo_path)
                sheet = workbook.active

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    usuario, estatus, rut, nombre, direccion, email, telefono, tipo_plan, tipo_cambio = row

                    # Validar los datos aquí
                    if not es_email_valido(email):
                        flash("El correo electrónico ingresado no es válido.", "danger")
                        return redirect(request.url)
                    if not es_rut_valido(rut):
                        flash("El RUT ingresado no es válido.", "danger")
                        return redirect(request.url)

                    cursor = mysql.connection.cursor()
                    cursor.execute(
                        "INSERT INTO clientes (usuario, estatus, rut, nombre, direccion, email, telefono, tipo_plan, tipo_cambio) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)",
                        (usuario, estatus, rut, nombre, direccion, email, telefono, tipo_plan, tipo_cambio)
                    )
                    mysql.connection.commit()
                    cursor.close()

                flash('Datos cargados exitosamente desde el archivo Excel', 'success')
                return redirect(url_for('lista_clientes'))

            except Exception as e:
                flash(f'Error al cargar el archivo Excel: {str(e)}', 'danger')

    return render_template('cargar_excel.html')

if __name__ == '__main__':
    app.run(debug=True, port=8080)



